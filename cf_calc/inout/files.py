import os
import pickle

from pprint import pformat
from collections import OrderedDict

import pandas as pd
import numpy as np

import openpyxl as oxl


max_excel_rows = 800000 # actual limit: 1048576
max_excel_cols = 10000 # actual limit: 16384


def add_full_directory(str):
    if str[1] == ':':  # string does begin with drive letter and :\
        return str

    else:
        if str[0] == '\\':  # string does begin with a sep
            return os.path.realpath('.') + str
        else:
            return os.path.realpath('.') + os.sep + str


def make_filepathext(file, path='', ext=''):
    # dealing with strings, which are immutable

    use_file = file
    use_path = path
    use_ext = ext

    # check the extension starts with '.'
    if len(use_ext) > 0:
        if use_ext[0] != '.':
            use_ext = '.' + use_ext

    # check the path ends with separator
    if len(use_path) > 0:
        use_path = add_full_directory(use_path)
        if use_path[-1] != os.sep:
            use_path = use_path + os.sep
    else:
        use_file = add_full_directory(file)

    #print(f'check_filepath output = {use_path + file + use_ext}')

    return use_path + use_file + use_ext


def get_write_mode(file, path='', ext=''):
    if os.path.exists(make_filepathext(file=file, path=path, ext=ext)):
        return 'a'  # append
    else:
        return 'w'  # write


def read_or_save_pickle(action, file, path='', ext='', list_save_vars=''):
    fullpathname = make_filepathext(file=file, path=path, ext=ext)

    if action == 'save':
        with open(fullpathname, 'wb') as f:
            pickle.dump(list_save_vars, f)

    elif action == 'read':
        with open(fullpathname, 'rb') as f:
            unpickled = pickle.load(f)

        return unpickled

    else:
        # function not called correctly
        raise TypeError(f'Function <read_or_save_pickle> called without '
                        f'"read" or "save" parameter')


def save_df_or_array(data, path, filename, extension):
    # save a dataframe or arrayt to a file

    bln_is_dataframe = isinstance(data, pd.core.frame.DataFrame)
    bln_is_nparray = isinstance(data, np.ndarray)

    fullpathname = make_filepathext(file=filename, path=path, ext=extension)

    if not (bln_is_dataframe or bln_is_nparray):
        raise KeyError(
            f'Function <save_df_or_array> was called with an unknown data type.  '
            f'We can handle pandas dataframe or numpy array.  '
            f'This was Data type = {type(data)}')

    if extension == '.csv':
        if bln_is_dataframe:
            data.to_csv(path_or_buf=fullpathname,
                        sep=',')
        elif bln_is_nparray:
            np.savetxt(fname=fullpathname,
                       X=data, delimiter=',')
        else:
            # shouldn't be able to get here; we already checked for type
            pass

    elif extension == '.npy':
        if bln_is_dataframe:
            np.save(file=fullpathname, arr=data.to_numpy())
        elif bln_is_nparray:
            np.save(file=fullpathname, arr=data)
        else:
            # shouldn't be able to get here; we already checked for type
            pass

    elif extension == '.xlsx':
        if bln_is_dataframe:
            # if writing to existing file...
            # with pd.ExcelWriter(fullpathname, mode='w') as writer:
            #     data.to_excel(writer, sheet_name='pandas')
            data.to_excel(fullpathname,
                          sheet_name=filename, index=True, header=True)

        elif bln_is_nparray:
            data = pd.DataFrame(data)
            data.to_excel(excel_writer=fullpathname,
                          sheet_name=filename, index=True, header=True)
        else:
            # shouldn't be able to get here; we already checked for type
            pass
    else:
        raise KeyError(
            f'function <save_df_or_array> was called with an unknown extension.  '
            f'File = {filename + extension} in directory = {path}')


def get_files(theDirectory, theString, theExtension):
    """
    Return two lists (full file names, and just file names), in all subdirectories of
    # theDirectory which match a string, and have an certain extension

    :param theDirectory:
    :param theString:
    :param theExtension:
    :return:
    """
    # region Source idea for finding files
    # # https://thispointer.com/python-how-to-get-list-of-files-in-directory-and-sub-directories/

    # listOfFiles = list()
    # for (dirpath, dirnames, filenames) in os.walk(dirName):
    #     listOfFiles += [os.path.join(dirpath, file) for file in filenames]
    # listOfFiles
    #
    # listOfFiles = list()
    # for (dirpath, dirnames, filenames) in os.walk(dirName):
    #     listOfFiles += [os.path.join(dirpath, file) for file in filenames if '.tif' in file]
    # listOfFiles
    # endregion

    listOfFiles = list()
    listOfFullFiles = list()
    for (dirpath, dirnames, filenames) in os.walk(theDirectory):
        listOfFullFiles += [os.path.join(dirpath, file) for file in filenames if
                            file.endswith(theExtension) and theString in file]
        listOfFiles += [file for file in filenames if
                        file.endswith(theExtension) and theString in file]
    return listOfFullFiles, listOfFiles


def dict_to_text(dictionary):
    # pprint was a bit finicky... if width is too wide, it seemed that
    # some nested levels showed up on the same line

    return pformat(dictionary, width=180, compact=False)


def write_dictionary_to_txt(dictionary, fname):
    str_rep = dict_to_text(dictionary=dictionary)

    with open(fname + '.txt', "w") as text_file:
        text_file.write(str_rep)


def write_df_to_excel(df, file, path='', ext='', sheet_name=''):
    fullpathname = make_filepathext(file=file, path=path, ext=ext)
    writemode = get_write_mode(file=fullpathname)
    print (f'fullpath = {fullpathname}, writemode = {writemode}')

    if sheet_name == '':
        sheet_name = 'df'

    if df.shape[0] > max_excel_rows or df.shape[1] > max_excel_cols:
        tempdf = df.iloc[0:12,0:12]
        tempdf.iloc[0,0] = f'df size = {df.shape[0]} x {df.shape[1]}... so not reported'
        with pd.ExcelWriter(path=fullpathname, mode=writemode) as writer:
            tempdf.to_excel(writer, sheet_name=sheet_name)
    else:
        with pd.ExcelWriter(path=fullpathname, mode=writemode) as writer:
            df.to_excel(writer, sheet_name=sheet_name)


def write_dict_to_excel(dict, file, path='', ext='', sheet_name=''):
    str_rep = dict_to_text(dictionary=dict)

    write_text_to_excel(thetext=str_rep, file_name=file, path=path,
                        ext=ext, sheet_name=sheet_name)


def write_text_to_excel(thetext, file_name, path='', ext='', sheet_name=''):
    # dump it all into one cell

    fullpathname = make_filepathext(file=file_name, path=path, ext=ext)
    writemode = get_write_mode(file=fullpathname)

    #print(f'write_text_to_excel fullpathname={fullpathname}')

    if writemode == 'a':
        wb = oxl.load_workbook(filename=fullpathname)
        if sheet_name != '':
            if sheet_name in wb.worksheets:
                print('duplicate sheet... openpyxl will rename')
    else:
        wb = oxl.Workbook()

        if sheet_name == '':
            sheet_name = 'Sheet'

    ws = wb.create_sheet(title=sheet_name)

    if thetext.count('\n') > max_excel_rows:
        stop_row = 12
    else:
        stop_row = max_excel_rows

    row_count = 1
    for line in thetext.splitlines():
        ws.cell(row_count, column=1).value = line
        row_count += 1
        if row_count > stop_row:
            ws.cell(row_count, column=1).value = f'too many lines, so not reported'
            break

    # ws['A1'].value = thetext
    # ws['A1'].alignment = oxl.styles.Alignment(wrap_text=True)
    # ws.column_dimensions['A'].width = 80

    # https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
    # for column_cells in ws.columns:
    #     length = max(len(as_text(cell.value)) for cell in column_cells)
    #     ws.column_dimensions[column_cells[0].column].width = length

    wb.save(filename=fullpathname)


def as_text(value):
    if value is None:
        return ""
    return str(value)


def split_dict(d):
    if isinstance(d, OrderedDict):
        d_strings = OrderedDict()
        d_dfs = OrderedDict()
    elif isinstance(d,dict):
        d_strings = {}
        d_dfs = {}
    else:
        raise TypeError('Function <split_dict> called without an '
                        'ordered or regular dictionary')

    for key,value in d.items():
        if isinstance(value, str):
            d_strings[key] = value
        else:
            d_dfs[key] = value

    return d_strings, d_dfs


#TODO: this should be recursive....
def write_geototpath_dict_to_excel(dict_gtp, file, path='', ext=''):

    fullname = make_filepathext(file=file, path=path, ext=ext)
    print(fullname)

    for geotot_key, value in dict_gtp.items():
        print(geotot_key)

        ds, ddfs = split_dict(value)
        write_dict_to_excel(ds,fullname,sheet_name=geotot_key+'_dict')

        for df_key, df in ddfs.items():
            if isinstance(df, pd.DataFrame):
                write_df_to_excel(df, file=fullname,sheet_name=geotot_key+'_df_'+df_key)


